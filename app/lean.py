from flask import Flask, abort, Response
from pathlib import Path;
import openpyxl
import json

DATAFOLDER = Path("./data")
DATASOURCES = [source for source in DATAFOLDER.iterdir() if source.is_file()]

#Utility Funktionen zum finden und aufbreiten von Datenquellen
def findDatasource(sourcename):
    return [datasource for datasource in DATASOURCES if datasource.name == ( sourcename + ".xlsx")]

def datasourceFullname(datasource):
    return "/".join([str(DATAFOLDER.parent), str(datasource)])

def doWithDatasource(sourcename, doThisWith):
    datasource = findDatasource(sourcename)
    if (len(datasource) > 0):
        return doThisWith(datasource[0])
    else:
        return abort(Response(json.dumps({'message': 'Datenquelle nicht gefunden'}), 404))

#Schlank gehaltene Cache Variante
#Da der Anwendungsfall sehr spezifisch ist und schnell umgesetzt werden muss
#Wird ein Server-Session basierter Cache genutzt der beim Serverneustart
#zurückgesetzt wird
cache = {}

#Key: {Datenquellenpfad}{Datenblatt}{Filterwert}
def keygen(file, sheetname, filterby):
    return datasourceFullname(file) + str(sheetname) + str(filterby)

def simplecache_find(file, sheetname, filterby):
    key = keygen(file, sheetname, filterby)
    if (key in cache):
        return { "found": True, "response": cache[key]}
    else:
        return { "found": False, "response": {}}

def simplecache_add(file, sheetname, filterby, response):
    key = keygen(file, sheetname, filterby)
    if (key not in cache):
        cache[key] = response

def cache_this(file, sheetname, filterby, fetchdata):
    cacheresult = simplecache_find(file, sheetname, filterby)
    if (cacheresult["found"]):
        return cacheresult["response"]
    response = fetchdata()
    simplecache_add(file, sheetname, filterby, response)

    return response

#Extrahiert die Jahre für die Datensätze vorhanden sind aus der Datenquelle
#Die Datenblätter sind alle nach dem gleichen Muster aufgebaut
#In Zeile 5, ab der neunten Spalte (I) findet sich pro Zelle ein Jahr wieder
#iter_cols liefert pro Eintrag alle Zellen die Daten beinhalten
DATASOURCE_YEAR_START_COL_INDEX = 9
DATASOURCE_YEAR_ROW_INDEX = 5
def bipyears(sheet):
    return [str(year[0].value) for year in sheet.iter_cols(
        min_col=DATASOURCE_YEAR_START_COL_INDEX,
        min_row=DATASOURCE_YEAR_ROW_INDEX,
        max_row=DATASOURCE_YEAR_ROW_INDEX
    )]

#Flache implementation, soll die Lesbarkeit erhöhen für Funktionen, die die Daten benötigen
DATASOURCE_DATA_START_COL_INDEX = 5
DATASOURCE_DATA_START_ROW_INDEX = 7
def bipdatabyrow(sheet):
    return sheet.iter_rows(
        min_col=DATASOURCE_DATA_START_COL_INDEX,
        min_row=DATASOURCE_DATA_START_ROW_INDEX
    )

#Tiefe Funktion zum extrahieren der Daten in der BIP-Datenquelle nach Datenblatt
#Funktioniert nur in den Datenblättern die die gewünschten Datenenthält
DATAROW_DATACELL_START_COL_INDEX = 3
DATALAYER = ["Bundesland", "Regierungsbezirke", "Kreisebene"]
FILTERVALUES = ["1", "2", "3"]
def extractbip(datasource, sheetname, filterby):
    book = openpyxl.load_workbook(datasourceFullname(datasource))
    sheet = book[sheetname]
    years = bipyears(sheet)
    yearsindexmap = [{ "index": index, "year": year } for index, year in enumerate(years)]
    dataraw = [datarow[DATAROW_DATACELL_START_COL_INDEX:] for datarow in bipdatabyrow(sheet) if str(datarow[filterby].value) == FILTERVALUES[filterby]]
    return {
        "ebene": DATALAYER[filterby],
        "jahre": years,
        "datensaetze": [{
            "gebietseinheit": dataset[0].value,
            "data": [{
                "jahr": yearindex["year"],
                "value": dataset[yearindex["index"] + 1].value
            } for yearindex in yearsindexmap]
        } for dataset in dataraw]
    }

def bipdata(datasource, sheetname, filterby=0):
    return json.dumps(cache_this(datasource, sheetname, filterby, lambda: extractbip(datasource, sheetname, filterby)))

#Konstanten benutzt um die Felder anzusteuern in den sich die entsprechenden Metadaten befinden
DATASOURCE_IMPRESSUM_COL_INDEX = 1
DATASOURCE_IMPRESSUM_START_INDEX = 1
DATASOURCE_IMPRESSUM_END_INDEX = 37
DATASOURCE_IMPRESSUM_TITLE_ROW_INDEX = 3
DATASOURCE_IMPRESSUM_HERAUSGABE_ROW_INDEX = 6
DATASOURCE_IMPRESSUM_URL_ROW_INDEX = 8
DATASOURCE_IMPRESSUM_ADRESSE_ROW_INDEX = 11
DATASOURCE_IMPRESSUM_ERSCHEINUNGSFOLGE_ROW_INDEX = 13
DATASOURCE_IMPRESSUM_ERSCHIENENAM_ROW_INDEX = 14
DATASOURCE_IMPRESSUM_BERECHNUNGSSTAND_ROW_INDEX = 16
DATASOURCE_IMPRESSUM_HISTORIE_START_ROW_INDEX = 17
DATASOURCE_IMPRESSUM_HISTORIE_ROW_COUNT = 8
DATASOURCE_IMPRESSUM_HINWEIS_START_ROW_INDEX = 34
DATASOURCE_IMPRESSUM_HINWEIS_ROW_COUNT = 3

#Flache implementation, soll die Lesbarkeit erhöhen für Funktionen, die die Daten benötigen
def impressumdata(datarows, index):
    return str(datarows[index].value)

#Funktion zum extrahieren des Impressums aus der BIP-Datenquelle
def extractimpressum(datasource):
    book = openpyxl.load_workbook(datasourceFullname(datasource))
    sheet = book["Impressum"]
    datarows = [datarow[0] for datarow in sheet.iter_rows(
        min_col=DATASOURCE_IMPRESSUM_COL_INDEX,
        min_row=DATASOURCE_IMPRESSUM_START_INDEX,
        max_row=DATASOURCE_IMPRESSUM_END_INDEX,
    )]
    return {
        "title": impressumdata(datarows, DATASOURCE_IMPRESSUM_TITLE_ROW_INDEX),
        "herausgabe": impressumdata(datarows, DATASOURCE_IMPRESSUM_HERAUSGABE_ROW_INDEX),
        "url": impressumdata(datarows, DATASOURCE_IMPRESSUM_URL_ROW_INDEX),
        "adresse": impressumdata(datarows, DATASOURCE_IMPRESSUM_ADRESSE_ROW_INDEX),
        "erscheinungsfolge": impressumdata(datarows, DATASOURCE_IMPRESSUM_ERSCHEINUNGSFOLGE_ROW_INDEX),
        "erschienenam": impressumdata(datarows, DATASOURCE_IMPRESSUM_ERSCHIENENAM_ROW_INDEX),
        "berechnungsstand": impressumdata(datarows, DATASOURCE_IMPRESSUM_BERECHNUNGSSTAND_ROW_INDEX),
        "historie": [impressumdata(datarows, DATASOURCE_IMPRESSUM_HISTORIE_START_ROW_INDEX + entryindex) for entryindex in range(DATASOURCE_IMPRESSUM_HISTORIE_ROW_COUNT)],
        "hinweise": [impressumdata(datarows, DATASOURCE_IMPRESSUM_HINWEIS_START_ROW_INDEX + entryindex) for entryindex in range(DATASOURCE_IMPRESSUM_HINWEIS_ROW_COUNT)]
    }

def impressum(datasource):
    return json.dumps(cache_this(datasource, "impressum", "0", lambda: extractimpressum(datasource)))

#Routing sections
app = Flask(__name__)

@app.route("/bip/<sourcename>/<sheetname>/bundeslaender")
def sheet_bundesland_details(sourcename, sheetname):
    return doWithDatasource(sourcename, lambda datasource: bipdata(datasource, sheetname, 0))

@app.route("/bip/<sourcename>/<sheetname>/regierungsbezirke")
def sheet_regierungsbezirke_details(sourcename, sheetname):
    return doWithDatasource(sourcename, lambda datasource: bipdata(datasource, sheetname, 1))

@app.route("/bip/<sourcename>/<sheetname>/kreisebene")
def sheet_kreisebene_details(sourcename, sheetname):
    return doWithDatasource(sourcename, lambda datasource: bipdata(datasource, sheetname, 2))

@app.route("/bip/<sourcename>/impressum")
def datasource_impressum(sourcename):
    return doWithDatasource(sourcename, lambda datasource: impressum(datasource))

if __name__ == '__main__':
	app.run(host='0.0.0.0', port=5000)