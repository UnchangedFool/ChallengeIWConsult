from flask import Flask
from pathlib import Path;
import openpyxl
import json

class Datasources:
    _DATAFOLDER = Path("./data")

    def __init__(self):
        self._DATASOURCES = [source for source in self._DATAFOLDER.iterdir() if source.is_file()]

    def findByName(self, sourcename):
        datasource = [source for source in self._DATASOURCES if source.name == ( sourcename + ".xlsx")]
        return BIPDatasource(datasource[0]) if len(datasource) > 0 else None

class Datasource:
    _datafolder = Path("./data")
    _source = ""

    def __init__(self, source):
        self._source = source

    def fullname(self):
        return "/".join([str(self._datafolder.parent), str(self._source)])

    def data(self, ident, filter):
        return None

    def impressum(self, ident):
        return None

#Schlank gehaltene Cache Variante
#Da der Anwendungsfall sehr spezifisch ist und schnell umgesetzt werden muss
#Wird ein Server-Session basierter Cache genutzt der beim Serverneustart
#zurückgesetzt wird
#Key: {Datenquellenpfad}{Datenblatt}{Filterwert}
class SimpleCache:
    _cache = {}

    def __init__(self):
        pass

    def _key(self, datasource, ident, filter):
        return datasource.fullname() + str(ident) + str(filter)

    def find(self, datasource, ident, filter):
        key = self._key(datasource, ident, filter)
        if (key in self._cache):
            return { "found": True, "response": self._cache[key]}
        else:
            return { "found": False, "response": {}}

    def add(self, datasource, ident, filter, response):
        key = self._key(datasource, ident, filter)
        if (key not in self._cache):
           self._cache[key] = response

    def cache_this(self, datasource, ident, filter, fetchdata):
        cacheresult = self.find(datasource, ident, filter)
        if (cacheresult["found"]):
            return cacheresult["response"]
        response = fetchdata()
        self.add(datasource, ident, filter, response)

        return response

#Extrahiert die Jahre für die Datensätze vorhanden sind aus der Datenquelle
#Die Datenblätter sind alle nach dem gleichen Muster aufgebaut
#In Zeile 5, ab der neunten Spalte (I) findet sich pro Zelle ein Jahr wieder
class BIPDatasource(Datasource):
    _DATASOURCE_YEAR_START_COL_INDEX = 9
    _DATASOURCE_YEAR_ROW_INDEX = 5

    def __init__(self, source):
        super().__init__(source)
        self._book = openpyxl.load_workbook(self.fullname())

    _DATASOURCE_DATA_START_COL_INDEX = 5
    _DATASOURCE_DATA_START_ROW_INDEX = 7
    def _databyrow(self, sheet):
        return sheet.iter_rows(
            min_col=self._DATASOURCE_DATA_START_COL_INDEX,
            min_row=self._DATASOURCE_DATA_START_ROW_INDEX
        )

    def years(self, sheetname):
        sheet = self._book[sheetname]
        return [str(year[0].value) for year in sheet.iter_cols(
            min_col=self._DATASOURCE_YEAR_START_COL_INDEX,
            min_row=self._DATASOURCE_YEAR_ROW_INDEX,
            max_row=self._DATASOURCE_YEAR_ROW_INDEX
        )]

    _DATAROW_DATACELL_START_COL_INDEX = 3
    _DATALAYER = ["Bundesland", "Regierungsbezirke", "Kreisebene"]
    _FILTERVALUES = ["1", "2", "3"]
    def _extractdata(self, sheetname, filterby):
        sheet = self._book[sheetname]
        years = self.years(sheetname)
        yearsindexmap = [{ "index": index, "year": year } for index, year in enumerate(years)]
        dataraw = [datarow[self._DATAROW_DATACELL_START_COL_INDEX:] for datarow in self._databyrow(sheet) if str(datarow[filterby].value) == self._FILTERVALUES[filterby]]
        return {
            "ebene": self._DATALAYER[filterby],
            "jahre": years,
            "datensaetze": [{
                "gebietseinheit": dataset[0].value,
                "data": [{
                    "jahr": yearindex["year"],
                    "value": dataset[yearindex["index"] + 1].value
                } for yearindex in yearsindexmap]
            } for dataset in dataraw]
        }

    def data(self, ident, filter):
        return json.dumps(SimpleCache().cache_this(self, ident, filter, lambda: self._extractdata(ident, filter)))

    #Konstanten benutzt um die Felder anzusteuern in den sich die entsprechenden Metadaten befinden
    _DATASOURCE_IMPRESSUM_COL_INDEX = 1
    _DATASOURCE_IMPRESSUM_START_INDEX = 1
    _DATASOURCE_IMPRESSUM_END_INDEX = 37
    _DATASOURCE_IMPRESSUM_TITLE_ROW_INDEX = 3
    _DATASOURCE_IMPRESSUM_HERAUSGABE_ROW_INDEX = 6
    _DATASOURCE_IMPRESSUM_URL_ROW_INDEX = 8
    _DATASOURCE_IMPRESSUM_ADRESSE_ROW_INDEX = 11
    _DATASOURCE_IMPRESSUM_ERSCHEINUNGSFOLGE_ROW_INDEX = 13
    _DATASOURCE_IMPRESSUM_ERSCHIENENAM_ROW_INDEX = 14
    _DATASOURCE_IMPRESSUM_BERECHNUNGSSTAND_ROW_INDEX = 16
    _DATASOURCE_IMPRESSUM_HISTORIE_START_ROW_INDEX = 17
    _DATASOURCE_IMPRESSUM_HISTORIE_ROW_COUNT = 8
    _DATASOURCE_IMPRESSUM_HINWEIS_START_ROW_INDEX = 34
    _DATASOURCE_IMPRESSUM_HINWEIS_ROW_COUNT = 3
    #Flache implementation, soll die Lesbarkeit erhöhen für Funktionen, die die Daten benötigen
    def _impressumcellvalue(self, datarows, index):
        return str(datarows[index].value)

    #Funktion zum extrahieren des Impressums aus der BIP-Datenquelle
    def _extractimpressum(self):
        sheet = self._book["Impressum"]
        datarows = [datarow[0] for datarow in sheet.iter_rows(
            min_col=self._DATASOURCE_IMPRESSUM_COL_INDEX,
            min_row=self._DATASOURCE_IMPRESSUM_START_INDEX,
            max_row=self._DATASOURCE_IMPRESSUM_END_INDEX,
        )]
        return {
            "title": self._impressumcellvalue(datarows, self._DATASOURCE_IMPRESSUM_TITLE_ROW_INDEX),
            "herausgabe": self._impressumcellvalue(datarows, self._DATASOURCE_IMPRESSUM_HERAUSGABE_ROW_INDEX),
            "url": self._impressumcellvalue(datarows, self._DATASOURCE_IMPRESSUM_URL_ROW_INDEX),
            "adresse": self._impressumcellvalue(datarows, self._DATASOURCE_IMPRESSUM_ADRESSE_ROW_INDEX),
            "erscheinungsfolge": self._impressumcellvalue(datarows, self._DATASOURCE_IMPRESSUM_ERSCHEINUNGSFOLGE_ROW_INDEX),
            "erschienenam": self._impressumcellvalue(datarows, self._DATASOURCE_IMPRESSUM_ERSCHIENENAM_ROW_INDEX),
            "berechnungsstand": self._impressumcellvalue(datarows, self._DATASOURCE_IMPRESSUM_BERECHNUNGSSTAND_ROW_INDEX),
            "historie": [self._impressumcellvalue(datarows, self._DATASOURCE_IMPRESSUM_HISTORIE_START_ROW_INDEX + entryindex) for entryindex in range(self._DATASOURCE_IMPRESSUM_HISTORIE_ROW_COUNT)],
            "hinweise": [self._impressumcellvalue(datarows, self._DATASOURCE_IMPRESSUM_HINWEIS_START_ROW_INDEX + entryindex) for entryindex in range(self._DATASOURCE_IMPRESSUM_HINWEIS_ROW_COUNT)]
        }

    def impressum(self):
        return json.dumps(SimpleCache().cache_this(self, "impressum", "0", lambda: self._extractimpressum()))

app = Flask(__name__)

@app.route("/bip/<sourcename>/<sheetname>/bundeslaender")
def sheet_bundesland_details(sourcename, sheetname):
    return Datasources().findByName(sourcename).data(sheetname, 0)

@app.route("/bip/<sourcename>/<sheetname>/regierungsbezirke")
def sheet_regierungsbezirke_details(sourcename, sheetname):
    return Datasources().findByName(sourcename).data(sheetname, 1)

@app.route("/bip/<sourcename>/<sheetname>/kreisebene")
def sheet_kreisebene_details(sourcename, sheetname):
    return Datasources().findByName(sourcename).data(sheetname, 2)

@app.route("/bip/<sourcename>/impressum")
def datasource_impressum(sourcename):
    return Datasources().findByName(sourcename).impressum()

if __name__ == '__main__':
	app.run(host='0.0.0.0', port=5000)