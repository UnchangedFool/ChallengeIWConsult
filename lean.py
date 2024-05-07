from flask import Flask
from pathlib import Path;
import openpyxl
import json

datafolder = Path("./data")
files = [x for x in datafolder.iterdir() if x.is_file()]
filtervalues = ["1", "2", "3"]
layer = ["Bundesland", "Regierungsbezirke", "Kreisebene"]

def getxlfile(filename):
    return [x for x in datafolder.iterdir() if x.name == ( filename + ".xlsx")]

def xlfull(file):
    return "/".join([str(datafolder.parent), str(file)])

def onxlfile(filename, func):
    xlfile = getxlfile(filename)
    if (len(xlfile) > 0):
        return func(xlfile[0])
    else:
        return "<h1>Datasource not found!</h1>"

#Schlank gehaltene Cache Variante
#Da der Anwendungsfall sehr spezifisch ist und schnell umgesetzt werden muss
#Wird ein Server-Session basierter Cache genutzt der beim Serverneustart
#zurückgesetzt wird
#Key: {Datenquellenpfad}{Datenblatt}{Filterwert}
cache = {}
def simplecache_find(file, sheetname, filterby):
    key = str(xlfull(file)) + str(sheetname) + str(filterby)
    if (key in cache):
        return { "found": True, "response": cache[key]}
    else:
        return { "found": False, "response": {}}

def simplecache_add(file, sheetname, filterby, response):
    key = str(xlfull(file)) + str(sheetname) + str(filterby)
    if (key not in cache):
        cache[key] = response

def cache_this(file, sheetname, filterby, fetchdata):
    cacheresult = simplecache_find(file, sheetname, filterby)

    if (cacheresult["found"]):
        return cacheresult["response"]

    response = fetchdata()

    simplecache_add(file, sheetname, filterby, response)

    return response

app = Flask(__name__)

@app.route("/data/<filename>")
def data_src_exists(filename):
    xlfile = getxlfile(filename)
    if (len(xlfile) > 0):
        return xlfull(xlfile[0])
    else:
        return "<h1>Datasource not found!</h1>"

#Extrahiert die Jahre für die Datensätze vorhanden sind aus der Datenquelle
#Die Datenblätter sind alle nach dem gleichen Muster aufgebaut
#In Zeile 5, ab der neunten Spalte (I) findet sich pro Zelle ein Jahr wieder
def years_from_datasheet(sheet):
    return [str(year[0].value) for year in sheet.iter_cols(min_col=9, min_row=5, max_row=5)]
#

def extractbip(file, sheetname, filterby):
    book = openpyxl.load_workbook(xlfull(file))
    sheet = book[sheetname]
    years = years_from_datasheet(sheet)
    yearsindexmap = [{ "index": index, "year": year } for index, year in enumerate(years)]
    dataraw = [row[3:] for row in sheet.iter_rows(min_col=5, min_row=7) if str(row[filterby].value) == filtervalues[filterby]]
    return {
        "ebene": layer[filterby],
        "jahre": years,
        "datensaetze": [{
            "gebietseinheit": dataset[0].value,
            "data": [{
                "jahr": yearindex["year"],
                "value": dataset[yearindex["index"] + 1].value
            } for yearindex in yearsindexmap]
        } for dataset in dataraw]
    }

def bipdata(file, sheetname, filterby=0):
    return json.dumps(cache_this(file, sheetname, filterby, lambda: extractbip(file, sheetname, filterby)))

@app.route("/bip/<filename>/<sheetname>/bundeslaender")
def sheet_bundesland_details(filename, sheetname):
    return onxlfile(filename, lambda f: bipdata(f, sheetname, 0))

@app.route("/bip/<filename>/<sheetname>/regierungsbezirke")
def sheet_regierungsbezirke_details(filename, sheetname):
    return onxlfile(filename, lambda f: bipdata(f, sheetname, 1))

@app.route("/bip/<filename>/<sheetname>/kreisebene")
def sheet_kreisebene_details(filename, sheetname):
    return onxlfile(filename, lambda f: bipdata(f, sheetname, 2))